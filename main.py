import openpyxl as xl
from datetime import datetime


def getDateFromCell(cell, format):
    text = cell.value
    # print(text)
    start_index = text.index('-')
    date = text[start_index + 1:]
    date = datetime.strptime(date, format)
    return date


def fill_black(ws, date_of_entering, all_end_date):
    start_col = 6
    black_fill = xl.styles.PatternFill(start_color='00000000',
                                       end_color='00000000',
                                       fill_type='solid')
    for y in all_end_date:
        if date_of_entering >= y:
            ws.cell(row=x, column=start_col).fill = black_fill
        start_col += 1


def fill_red(ws, top_length, row):
    should_fill_red = True
    red_fill = xl.styles.PatternFill(start_color='FFFF0000',
                                     end_color='FFFF0000',
                                     fill_type='solid')
    for i in range(top_length, 4, -1):
        if ws.cell(row=row, column=i).value:
            should_fill_red = False
            break
    if should_fill_red:
        ws.cell(row=row, column=1).fill = red_fill
        ws.cell(row=row, column=2).fill = red_fill


def should_kick(ws, top_length, row):
    should_kick = True
    for i in range(top_length, top_length - 8, -1):
        if ws.cell(row=row, column=i).value or (
                ws.cell(row=row, column=i).fill.start_color.rgb == '00000000' and not ws.cell(row=row,
                                                                                              column=i).fill.patternType is None):
            should_kick = False
            break
    if not should_kick:
        for i in range(top_length, 4, -1):
            if ws.cell(row=row, column=i).value:
                return False
    return True


def fill_yellow(ws, top_length, row):
    should_fill_yellow = True
    yellow_fill = xl.styles.PatternFill(start_color='FFFFC000',
                                        end_color='FFFFC000',
                                        fill_type='solid')
    for i in range(top_length, top_length - 5, -1):
        if ws.cell(row=row, column=i).value or (
                ws.cell(row=row, column=i).fill.start_color.rgb == '00000000' and not ws.cell(row=row,
                                                                                              column=i).fill.patternType is None):
            should_fill_yellow = False
            break
    if should_fill_yellow:
        ws.cell(row=row, column=1).fill = yellow_fill
        ws.cell(row=row, column=2).fill = yellow_fill


def mark_consecutive_four_goals_achieved(achievement_list):
    achieved_counter = 0
    total_allowed_counter = 0
    start_indexes = []
    for i in range(len(achievement_list)):
        if achievement_list[i] == '达标':
            achieved_counter += 1
        else:
            achieved_counter = 0
        if achieved_counter == 4 and total_allowed_counter < 2:
            start_indexes.append(i - 3)
            total_allowed_counter += 1
            achieved_counter = 0
    return start_indexes


def fill_green(ws, row, achievement_list):
    consecutive_four_indexes = mark_consecutive_four_goals_achieved(achievement_list)
    green_fill = xl.styles.PatternFill(start_color='FF92D050',
                                       end_color='FF92D050',
                                       fill_type='solid')
    if len(consecutive_four_indexes):
        ws.cell(row=row, column=1).fill = green_fill
        ws.cell(row=row, column=2).fill = green_fill
        for i in range(len(consecutive_four_indexes)):
            for j in range(4):
                ws.cell(row=row, column=6 + consecutive_four_indexes[i] + j).fill = green_fill

    pass



if __name__ == '__main__':
    wb = xl.load_workbook(filename="6k萌新群成员段位及课题完成情况 2019.12.9.xlsx")
    ws = wb.active

    date_format1 = '%Y.%m.%d'
    date_format2 = '%Y/%m/%d'

    qq_num_file = "qq_nums.txt"
    num_list = []
    should_del_row = []
    should_be_kicked = []

    try:
        with open(qq_num_file, "r+", encoding='utf-8') as f:
            for line in f:
                num_list.append(line.strip("\n"))
    except FileNotFoundError:
        print("qq num list file not found, continue...")
        pass

    all_end_date = []
    top_row = ws[1]
    top_length = len(top_row)
    row_count = ws.max_row
    for x in range(6, top_length + 1):
        all_end_date.append(getDateFromCell(ws.cell(row=1, column=x), date_format1))
    print(row_count)
    for x in range(2, row_count + 1):
        date_of_entering = ws['C%d' % x].value
        qq_num = ws['A%d' % x].value
        if len(num_list):
            if str(qq_num) not in num_list:
                should_del_row.append(x)
                print("(" + str(qq_num) + ")" + str(ws['B%d' % x].value) + " should be removed.")
                continue
        if not date_of_entering:
            break
        current_date = datetime.now()
        fill_black(ws, date_of_entering, all_end_date)
        achievement_list = [i.value for i in ws[x][5:]]
        fill_green(ws, x, achievement_list)
        fill_yellow(ws, top_length, x)
        if (current_date - date_of_entering).days <= 30:
            fill_red(ws, top_length, x)
        else:
            if should_kick(ws, top_length, x):
                should_be_kicked.append((ws.cell(row=x, column=1).value, ws.cell(row=x, column=2).value))
    print("Would you like to remove all users that do not exist? Press y to confirm delete or press anything else to move on.")
    x = input()
    if x in ["y", "Y"]:
        should_del_row.reverse()
        for row in should_del_row:
            ws.delete_rows(row)
    print("List of people should be kicked:")
    for item in should_be_kicked:
        print(item)
    wb.save('Sample.xlsx')
